'''
@Author: Nik Sauer 
'''
import argparse
import random
import string
import sys
import unicodedata
from typing import Generator, Tuple

import openpyxl
from openpyxl import load_workbook
import logging
import logging.handlers

import unidecode
import re

logger = logging.getLogger(__name__)
handler = logging.handlers.RotatingFileHandler("create_user.log", maxBytes=10000, backupCount=5)
formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
handler.setFormatter(formatter)

stream_handler = logging.StreamHandler(sys.stdout)
stream_handler.setFormatter(formatter)

logger.addHandler(handler)
logger.addHandler(stream_handler)

verbose = False
quiet = False

existing_user = []

def read_file(path: str) -> Generator:
    """
    Liest Excel datei und gibt zeilen aus (Yield)
    :param path:
    :return:
    """
    try:
        wb = load_workbook(path, read_only=True)
        ws = wb[wb.sheetnames[0]]
        for row in ws.iter_rows(min_row=2):
            if all(cell.value is None for cell in row):
                continue
            # firstname	lastname	group	class
            yield (row[0].value, row[1].value, row[2].value, row[3].value)
    except:
        logger.error("Datei nicht gefunden")

def shave_marks(s: str) -> str:
    """
    Remove all diacritic marks
    :param s:
    :return:
    """
    norm_txt = unicodedata.normalize('NFD', s)
    shaved = ''.join(c for c in norm_txt
                     if not unicodedata.combining(c))
    return unicodedata.normalize('NFC', shaved)


def replace_umlaut(s: str) -> str:
    """
    Ersetzt Umlaute
    :param s:
    :return:
    """
    replacements = {
        'ä': 'ae',
        'ö': 'oe',
        'ü': 'ue',
        'Ä': 'Ae',
        'Ö': 'Oe',
        'Ü': 'Ue',
        'ß': 'ss'
    }
    for umlaut, replacement in replacements.items():
        s = s.replace(umlaut, replacement)
    return shave_marks(s)

def check_name(name: str) -> str:
    """
    überprüft ob name passt
    :param name:
    :return:
    """
    name = unidecode.unidecode(name)

    name = replace_umlaut(name) 

    name = name.lower()
    name = name.replace(' ', '_')
    name = re.sub(r'[^a-z0-9_]', '', name)
    return name


def generate_unique_name(name: str) -> str:
    """
    Generiert einen eindeutigen namen
    :param name:
    :return:
    """
    global existing_user
    if name not in existing_user:
        existing_user.append(name)
        return name
    index = 1
    while f"{name}{index}" in existing_user:
        index += 1
    new_name = f"{name}{index}"
    existing_user.append(new_name)
    return new_name

def create_user(file, pwd: str, first_name: str, last_name: str, group: str, class_name: str) -> None:
    """
    UserCreates (Commands) erstellen
    :param file:
    :param pwd:
    :param first_name:
    :param last_name:
    :param group:
    :param class_name:
    :return:
    """
    logger.info(f"Creating user: {first_name}_{last_name}")
    if verbose:
        print(f"echo creating user: {first_name}_{last_name}", file=file)

    command1 = (
        f"getent passwd {first_name}_{last_name} > /dev/null && "
        f"echo 'User {first_name}_{last_name} already exists. Aborting.' && "
        f"exit 1 || true"
    )

    command2 = f"groupadd {last_name}"

    command3 = (
        f"useradd -d /home/{last_name} -c {last_name} "
        f"-m -g {last_name} -G {group},{class_name} "
        f"-s /bin/bash {first_name}_{last_name}"
    )
    command4 = f"echo {first_name}_{last_name}:{escape_quote(pwd)} | chpasswd"

    print(command1, file=file)
    print(command2, file=file)
    print(command3, file=file)
    print(command4, file=file)




def create_files(path: str) -> None:
    """
    Itariert durch alle User und erzeugt die dazugehörigen files
    :param path:
    :return:
    """
    logger.info("Starting file creation")
    worksheet, sheet = create_credentials()
    row = 2
    with open("create_user.sh", "w", encoding="UTF-8") as create_user_file, open("delete_user.sh", "w", encoding="UTF-8") as delete_user_file:
        print("#!/bin/bash", file=create_user_file)
        print("#!/bin/bash", file=delete_user_file)

        print("set -e", file=create_user_file)
        print("set -e", file=delete_user_file)

        print("mkdir /home", file=create_user_file)

        verwendetePWSs = set()

        for firstname, lastname, group, _class in read_file(path):
            first_name = check_name(str(firstname).lower())

            last_name = generate_unique_name(check_name(str(lastname).lower()))

            group = str(group).lower()
            class_name = str(_class)

            pwd = generate_password(12, verwendetePWSs)

            verwendetePWSs.add(pwd)
            create_user(create_user_file, pwd, first_name, last_name, group, class_name)
            delete_user(delete_user_file, first_name, last_name)
            add_credentials(sheet, row, pwd, last_name)

            row += 1
    save_credentials(worksheet)
    logger.info("Files created")




def main():
    global verbose, quiet

    parser = argparse.ArgumentParser()
    parser.add_argument("filename", type=str, help="filename")

    loglevel = parser.add_mutually_exclusive_group()
    loglevel.add_argument("-v", "--verbose", action="store_true", help="activates verbose mode")
    loglevel.add_argument("-q", "--quiet", action="store_true", help="activates quite mode")

    args = parser.parse_args()

    print(args.filename)

    path = args.filename
    verbose = args.verbose
    quiet = args.quiet

    configure_logging()

    try:
        create_files(path)
    except FileNotFoundError:
        logger.error("File not found")

if __name__ == "__main__":
    main()















